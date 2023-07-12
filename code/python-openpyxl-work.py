# ```python
import openpyxl

# Excelファイルを開く
wb = openpyxl.load_workbook('sample.xlsx')

# シートを選択する
ws = wb['Sheet1']

# 1行目の1列目から値を取得する
cell = ws.cell(row=1, column=1)
print(cell.value)
# ```


# ```python
import openpyxl

# Excelファイルを開く
wb = openpyxl.load_workbook('sample.xlsx')

# シートを選択する
ws = wb['Sheet1']

# 2行目の1列目に値を書き込む
ws.cell(row=2, column=1, value='Hello, World!')

# Excelファイルを保存する
wb.save('sample.xlsx')
# ```

